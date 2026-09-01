"""Three-sheet financial-summary Excel export.

Sheet 1 (Summary): headline figures, category totals and monthly cashflow trend.
Sheet 2 (Monthly Detail): the credit-team reconciliation table — Credits/Loans/
Outliers/Net per direction, plus the balance range, one row per month.
Sheet 3 (Transactions): every extracted transaction with its classification;
flagged rows are highlighted and always carry a Flag reason.

This is the reconciliation artefact the credit team asked for — and a usable
fallback when the scorecard PDF is slow or fails, because it's generated
independently from the same persisted data.
"""
from __future__ import annotations

from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.config import settings
from app.services.summary.financial_summary import breakdown_slices

# Font name only — Excel doesn't embed fonts the way a PDF must, so this is
# safe to set even on a machine without Consolas installed; Excel silently
# substitutes a fallback there instead of failing to render.
_FONT_NAME = "Consolas"
# Table column-header fill: amber with dark text, distinct from the purple
# brand accent used for titles — kept readable at any zoom level.
_HEADER_FILL = PatternFill("solid", fgColor="F2A93C")
_HEADER_FONT = Font(name=_FONT_NAME, color="1F1400", bold=True)
_TOTAL_FILL = PatternFill("solid", fgColor="E8E5F5")
_FLAGGED_FILL = PatternFill("solid", fgColor="FDF3E7")
_TITLE_FONT = Font(name=_FONT_NAME, bold=True, size=14, color="3B2F8F")
_SUBTITLE_FONT = Font(name=_FONT_NAME, size=9, color="6B7280")
_BOLD = Font(name=_FONT_NAME, bold=True)
_BODY_FONT = Font(name=_FONT_NAME)


def _style_header(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _apply_default_font(ws) -> None:
    """Consolas on every cell that doesn't already carry an explicit style."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.font is None or cell.font.name != _FONT_NAME:
                existing = cell.font
                cell.font = Font(
                    name=_FONT_NAME, bold=existing.bold, italic=existing.italic,
                    color=existing.color, size=existing.size or 11,
                )


def _autosize(ws, max_width: int = 60) -> None:
    for col in ws.columns:
        length = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(length + 2, max_width)


def _brand_header(ws, subtitle: str) -> int:
    """Writes the brand wordmark + subtitle, returns the next free row."""
    ws["A1"] = settings.app_name
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = subtitle
    ws["A2"].font = _SUBTITLE_FONT
    ws["A3"] = f"Generated {datetime.now(timezone.utc).strftime('%d %b %Y %H:%M UTC')}"
    ws["A3"].font = _SUBTITLE_FONT
    return 5


def build_financial_workbook(statement, score) -> str:
    summary = score.financial_summary or {}
    out_dir = settings.storage_path / "reports"
    path = str(out_dir / f"financial_summary_{statement.reference_id}.xlsx")

    wb = Workbook()

    # ---- Summary sheet ----
    ws = wb.active
    ws.title = "Summary"
    row = _brand_header(ws, "Transparent credit scoring, explained.")
    ws.cell(row=row, column=1, value=f"Reference: {statement.reference_id}")
    row += 1
    ws.cell(row=row, column=1,
            value=f"Client: {statement.account_holder or '—'}  |  National ID: {statement.national_id or '—'}  |  "
                  f"Account: {statement.account_number or '—'}")
    row += 1
    ws.cell(row=row, column=1,
            value=f"Credit score: {score.credit_score}  |  Grade: {score.grade}  |  "
                  f"Affordability: {score.limit_low:,.0f}–{score.limit_high:,.0f}  |  "
                  f"Needs review: {'Yes' if statement.needs_review else 'No'}")
    row += 2

    ws.cell(row=row, column=1, value="Metric")
    ws.cell(row=row, column=2, value="Value")
    _style_header(ws, row, 2)
    row += 1

    def put(label: str, value) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1

    totals = summary.get("totals", {})
    behaviour = summary.get("behaviour", {})
    lending = summary.get("lending", {})
    period = summary.get("period", {})
    fraud = score.fraud_data or {}
    ratios = (score.score_breakdown or {}).get("ratios", {})
    for label, value in [
        ("Debt to Income", ratios.get("debt_to_income")),
        ("Income Volatility", ratios.get("income_volatility")),
        ("Total received", totals.get("total_received", 0)),
        ("Total sent", totals.get("total_sent", 0)),
        ("Net position", totals.get("net_position", 0)),
        ("Transaction count", totals.get("transaction_count", 0)),
        ("Avg credit", totals.get("avg_credit", 0)),
        ("Avg debit", totals.get("avg_debit", 0)),
        ("Loan received total", lending.get("loan_received_total", 0)),
        ("Loan repaid total", lending.get("loan_repaid_total", 0)),
        ("Fuliza out", lending.get("fuliza_out", 0)),
        ("Betting out", behaviour.get("betting_out", 0)),
        ("Salary in", behaviour.get("salary_in", 0)),
        ("P2P received", behaviour.get("p2p_received", 0)),
        ("P2P sent", behaviour.get("p2p_sent", 0)),
        ("Contra (self-transfer) total", behaviour.get("contra_total", 0)),
        ("Contra (self-transfer) count", behaviour.get("contra_count", 0)),
        ("Outlier credit total", behaviour.get("outlier_credit_total", 0)),
        ("Outlier debit total", behaviour.get("outlier_debit_total", 0)),
        ("Expenses-to-income", behaviour.get("expenses_to_income", 0)),
        ("Statement months", period.get("statement_months", 0)),
        ("Active months", period.get("active_months", 0)),
        ("Fraud risk score", fraud.get("risk_score", 0)),
        ("Fraud risk level", fraud.get("risk_level", "—")),
    ]:
        put(label, value)

    # Category block.
    row += 1
    ws.cell(row=row, column=1, value="Category")
    ws.cell(row=row, column=2, value="In")
    ws.cell(row=row, column=3, value="Out")
    ws.cell(row=row, column=4, value="Count")
    _style_header(ws, row, 4)
    row += 1
    for cat, vals in (summary.get("categories") or {}).items():
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=vals.get("in", 0))
        ws.cell(row=row, column=3, value=vals.get("out", 0))
        ws.cell(row=row, column=4, value=vals.get("count", 0))
        row += 1

    # Breakdown block (Betting / Salary / Loans / Remittance / Other) — the
    # same curated slices the PDF's donut chart uses, so the two can never
    # show a different breakdown of the same statement. Charted natively so
    # it's a real, editable Excel pie chart, not a picture of one.
    slices = breakdown_slices(summary)
    if slices:
        row += 1
        breakdown_header_row = row
        ws.cell(row=row, column=1, value="Category")
        ws.cell(row=row, column=2, value="Amount")
        ws.cell(row=row, column=3, value="% of activity")
        _style_header(ws, row, 3)
        row += 1
        breakdown_first_data_row = row
        for s in slices:
            ws.cell(row=row, column=1, value=s["label"])
            ws.cell(row=row, column=2, value=s["value"])
            ws.cell(row=row, column=3, value=s["pct"])
            ws.cell(row=row, column=3).number_format = "0%"
            row += 1
        breakdown_last_data_row = row - 1

        chart = PieChart()
        chart.title = "Financial Breakdown"
        chart.height, chart.width = 7, 10
        data = Reference(ws, min_col=2, min_row=breakdown_header_row, max_row=breakdown_last_data_row)
        labels = Reference(ws, min_col=1, min_row=breakdown_first_data_row, max_row=breakdown_last_data_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        ws.add_chart(chart, f"F{breakdown_header_row}")

    # Monthly cashflow trend block (headline received/sent/balance only —
    # see the "Monthly Detail" sheet for the full Credits/Loans/Outliers/Net
    # reconciliation table).
    row += 1
    ws.cell(row=row, column=1, value="Month")
    ws.cell(row=row, column=2, value="Received")
    ws.cell(row=row, column=3, value="Sent")
    ws.cell(row=row, column=4, value="Closing balance")
    _style_header(ws, row, 4)
    row += 1
    trends = summary.get("trends", {})
    months = sorted(set(trends.get("received", {})) | set(trends.get("sent", {})) | set(trends.get("balance", {})))
    for m in months:
        ws.cell(row=row, column=1, value=m)
        ws.cell(row=row, column=2, value=trends.get("received", {}).get(m, 0))
        ws.cell(row=row, column=3, value=trends.get("sent", {}).get(m, 0))
        ws.cell(row=row, column=4, value=trends.get("balance", {}).get(m))
        row += 1

    _autosize(ws)

    # ---- Monthly Detail sheet (the credit-team reconciliation table) ----
    monthly = summary.get("monthly_detail") or {}
    if monthly.get("rows"):
        md = wb.create_sheet("Monthly Detail")
        row = _brand_header(md, "Monthly reconciliation — Credits / Loans / Outliers / Net, per direction")
        headers = ["Month", "Credits (Cr)", "Loans (Cr)", "Outliers (Cr)", "Net (CR)",
                   "Debits (Dr)", "Loans (Dr)", "Outliers (Dr)", "Net (DR)", "Highest Bal", "Lowest Bal"]
        header_row = row
        for c, h in enumerate(headers, start=1):
            md.cell(row=header_row, column=c, value=h)
        _style_header(md, header_row, len(headers))

        r = header_row + 1
        for mr in monthly["rows"]:
            md.cell(row=r, column=1, value=mr["month"])
            md.cell(row=r, column=2, value=mr["credits"])
            md.cell(row=r, column=3, value=mr["loan_credits"])
            md.cell(row=r, column=4, value=mr["outlier_credits"])
            md.cell(row=r, column=5, value=mr["net_credit"])
            md.cell(row=r, column=6, value=mr["debits"])
            md.cell(row=r, column=7, value=mr["loan_debits"])
            md.cell(row=r, column=8, value=mr["outlier_debits"])
            md.cell(row=r, column=9, value=mr["net_debit"])
            md.cell(row=r, column=10, value=mr["highest_balance"])
            md.cell(row=r, column=11, value=mr["lowest_balance"])
            r += 1

        for label, key_row in (("Total", monthly.get("totals", {})), ("Average", monthly.get("averages", {}))):
            md.cell(row=r, column=1, value=label).font = _BOLD
            for c, key in enumerate(("credits", "loan_credits", "outlier_credits", "net_credit",
                                      "debits", "loan_debits", "outlier_debits", "net_debit",
                                      "highest_balance", "lowest_balance"), start=2):
                cell = md.cell(row=r, column=c, value=key_row.get(key))
                cell.font = _BOLD
                cell.fill = _TOTAL_FILL
            r += 1

        contra_total = summary.get("behaviour", {}).get("contra_total", 0)
        contra_count = summary.get("behaviour", {}).get("contra_count", 0)
        if contra_count:
            r += 1
            md.cell(row=r, column=1,
                    value=f"Contra entries excluded from the above: {contra_count} self-transfer(s), "
                          f"totaling KSh {contra_total:,.2f}.")
        _autosize(md)

    # ---- Transactions sheet ----
    tx = wb.create_sheet("Transactions")
    headers = ["Date", "Reference", "Description", "Paid In", "Withdrawn", "Balance",
               "Label", "Category", "Flagged", "Flag reason"]
    tx.append(headers)
    _style_header(tx, 1, len(headers))
    n_flagged = 0
    for t in statement.transactions:
        tx.append([
            t.transaction_datetime.isoformat() if t.transaction_datetime else "",
            t.transaction_ref or "",
            t.description or "",
            t.paid_in,
            t.withdrawn,
            t.balance,
            t.label,
            t.category or "",
            "Yes" if t.is_flagged else "No",
            t.flag_reason or "",
        ])
        if t.is_flagged:
            n_flagged += 1
            for c in range(1, len(headers) + 1):
                tx.cell(row=tx.max_row, column=c).fill = _FLAGGED_FILL
    tx.freeze_panes = "A2"
    tx.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{tx.max_row}"
    _autosize(tx)

    for sheet in wb.worksheets:
        _apply_default_font(sheet)

    wb.save(path)
    return path
